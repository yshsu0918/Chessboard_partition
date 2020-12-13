import os
from chessboard_tool import *

# Function: csv 取得補丁資訊, 然後把字串代換.

# parameters
## input path
input_path = './chao70r/chao70_spin_oldmask_fixterritory_nolastpass_fixbyhuman_addtargetwithpudding'
xls_path = './chao70r/補丁1.xlsx'
## output path
output_path = './chao70r/chao70_spin_oldmask_fixterritory_nolastpass_fixbyhuman_addtargetwithpudding'
# end parameters

if not os.path.isdir(output_path) :
    os.mkdir(output_path)

from openpyxl import load_workbook

wb = load_workbook(xls_path)
ws = wb.active
print(wb.sheetnames)
for row in ws.rows:
    fn = row[0].value
    #print(row[1].value)
    if row[1].value == 'del':
        str2sgf( os.path.join(output_path,fn) , '')
        continue
    else:
        replaceA = row[1].value
        replaceB = row[2].value

    sgf_str = sgf2str( os.path.join(input_path, fn) )
    A = sgf_str.replace( replaceA, replaceB ).replace( ';W[];B[]', '' ).replace( ';B[];W[]', '' )
    str2sgf( os.path.join(output_path,fn) , A)
    

    chao70_spin_oldmask_fixterritory_nolastpass/